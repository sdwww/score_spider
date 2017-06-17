import os
import re
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

from ScoreThread import ScoreThread


# 保存网页内容
def save_html(save_path, file_name, content):
    try:
        if not os.path.isdir(save_path + '/'):
            os.mkdir(save_path + '/')
        html = open(save_path + '/' + file_name + '.html', 'w', encoding='utf-8')
        html.write(content)
        html.close()
        return True
    except IOError:
        return False


def get_school_dict(session, head):
    # 获取大学与id关系并保存
    school_dict = {}
    for i in range(1, 5):
        content = session.get(
            'http://data.api.gkcx.eol.cn/soudaxue/queryschool.html?messtype=jsonp&'
            'callback=jQuery18307013630659283934_1497580101517&province=&schooltype=&'
            'page=' + str(i) + '&size=30&keyWord1=&schoolprop=&'
                               'schoolflag=211%E5%B7%A5%E7%A8%8B&schoolsort=&schoolid=&_=1497580101995',
            headers=head).text
        page_name = 'school_id_' + str(i)
        if save_html('./saved_html', page_name, content):
            print('网页' + page_name + '保存成功')
        else:
            print('网页' + page_name + '保存失败')
        # print(content[content.find('schoolid'):150])

        p_id = r'schoolid.+"(.+)"'
        pattern_id = re.compile(p_id)
        school_id_list = pattern_id.findall(content)

        p_name = r'schoolname.+"(.+)"'
        pattern_name = re.compile(p_name)
        school_name_list = pattern_name.findall(content)

        for i in range(len(school_id_list)):
            school_dict[school_name_list[i]] = school_id_list[i]
    return school_dict


def get_province_dict():
    # 获取省份与id关系并保存
    province_dict = {"安徽": "10008", "澳门": "10145", "北京": "10003", "重庆": "10028", "福建": "10024",
                     "甘肃": "10023", "贵州": "10026", "广东": "10011", "广西": "10012", "河北": "10016",
                     "河南": "10017", "黑龙江": "10031", "湖北": "10021", "湖南": "10022", "海南": "10019",
                     "江苏": "10014", "江西": "10015", "吉林": "10004", "辽宁": "10027", "内蒙古": "10002",
                     "宁夏": "10007", "青海": "10030", "上海": "10000", "山东": "10009", "山西": "10010",
                     "陕西": "10029", "四川": "10005", "天津": "10006", "新疆": "10013", "西藏": "10025",
                     "香港": "10020", "云南": "10001", "台湾": "10146", "浙江": "10018"}
    return province_dict


# 生成路径
def build_path(school_dict, province_dict):
    subject_dict = {'文科': 10034, '理科': 10035}
    complete_path = []
    for province_name, province_id in province_dict.items():
        for subject_name, subject_id in subject_dict.items():
            for school_name, school_id in school_dict.items():
                for year in range(2014, 2017):
                    url = 'http://gkcx.eol.cn/schoolhtm/specialty/' + str(
                        school_id) + '/' + str(subject_id) + '/specialtyScoreDetail_' \
                          + str(year) + '_' + str(province_id) + '.htm'
                    complete_path.append([province_name, subject_name, school_name + str(year), url])
    return complete_path


# 创建多线程
def create_threads(thread_count, session, lock, path, head):
    threads = []
    for i in range(thread_count):
        score_thread = ScoreThread(name=i, session=session, lock=lock, path=path, head=head)
        score_thread.start()
        threads.append(score_thread)
    for thread in threads:
        thread.join()


# 读取错误文件
def read_error_log():
    error_log = []
    file_object = open('.\log\error.txt', 'r+', encoding='utf-8')
    try:
        all_the_text = file_object.readlines()
        for i in all_the_text:
            items = i.replace('\n', '').split(' ')
            error_log.append(items)
    finally:
        file_object.close()
        file_object = open('.\log\error.txt', 'r+', encoding='utf-8')
        file_object.truncate()
        file_object.close()
    return error_log


def parser_html(path):
    # wb = load_workbook(filename='./data_xlsx/SPM_COUNT.xlsx')
    # sheetnames = wb.get_sheet_names()
    # ws = wb.get_sheet_by_name(sheetnames[0])
    # delete_count = 0
    # ws.cell(column=4, row=rx, value=spm_drug_dict[i])
    # wb.save('./data_xlsx/SPM_COUNT.xlsx')
    try:
        if not os.path.isdir('./saved_excel/'):
            os.mkdir('./saved_excel/')
    except FileExistsError:
        pass
    for i in path:
        try:
            if not os.path.isdir('./saved_excel/' + i[0] + '/'):
                os.mkdir('./saved_excel/' + i[0] + '/')
            if not os.path.exists('./saved_excel/' + i[0] + '/' + i[0] + i[1] + '.xlsx'):
                wb = Workbook()
                ws = wb.active
                ws.append(['学校', '专业名称', '年份', '最高分', '平均分', '最低分', '录取批次'])
                wb.save('./saved_excel/' + i[0] + '/' + i[0] + i[1] + '.xlsx')
        except FileExistsError:
            pass
        soup = BeautifulSoup(open('./saved_html/' + i[0] + '/' + i[1] + '/' + i[2] + '.html',
                                  encoding="utf-8"), "lxml")
        all_row = soup.find('table').find('tbody').find_all('tr')
        wb = load_workbook('./saved_excel/' + i[0] + '/' + i[0] + i[1] + '.xlsx')
        is_null = False
        for one_row in all_row:
            row_value = [i[2][0:-4]]
            for item in one_row.find_all('td'):
                if item.string.replace('\t', '').replace(' ', '').replace('\n', '') == '暂时没有数据':
                    is_null = True
                else:
                    row_value.append(item.string.replace('\t', '').replace(' ', '').replace('\n', ''))
            if is_null:
                break
            sheetnames = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheetnames[0])
            ws.append(row_value)
        wb.save('./saved_excel/' + i[0] + '/' + i[0] + i[1] + '.xlsx')
        if not is_null:
            print('保存./saved_html/' + i[0] + '/' + i[1] + '/' + i[2] + '.html成功！')


if __name__ == "__main__":
    start = time.clock()
    score_session = requests.session()
    # 构造headers
    agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 ' \
            '(KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36'
    headers = {
        "Host": "gkcx.eol.cn",
        'User-Agent': agent
    }

    school_info = get_school_dict(score_session, headers)
    province_info = get_province_dict()
    complete_path = build_path(school_info, province_info)

    # score_lock = threading.RLock()
    # create_threads(thread_count=1, session=score_session, lock=score_lock, path=complete_path, head=headers)
    # error_list = read_error_log()
    # while len(error_list):
    #     score_lock = threading.RLock()
    #     create_threads(thread_count=1, session=score_session, lock=score_lock, path=error_list, head=headers)
    #     error_list = read_error_log()
    parser_html(complete_path)
    print('总时间为:', time.clock() - start)
